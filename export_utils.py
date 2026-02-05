from datetime import datetime

def export_excel(rows, save_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"
    headers = ["Article", "Nature", "Quantité", "Prix", "Seuil mini", "Date ajout", "Observation", "Valeur (Qté*Prix)"]
    ws.append(headers)
    # Styles et ajout des données
    header_fill = PatternFill("solid", fgColor="1976D2")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in rows:
        valeur = (r[2] or 0) * float(r[3] or 0)
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], valeur])
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = max(12, min(50, max_len+2))
    wb.save(save_path)

def export_pdf(rows, save_path, title="Inventaire Magasin"):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    doc = SimpleDocTemplate(save_path, pagesize=landscape(A4), topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#1976D2")
    elems = [Paragraph(title, title_style), Spacer(1, 12)]
    data = [["Article", "Nature", "Quantité", "Prix", "Seuil mini", "Date ajout", "Observation", "Valeur (Qté*Prix)"]]
    for r in rows:
        valeur = (r[2] or 0) * float(r[3] or 0)
        data.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], valeur])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1976D2")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor("#E3F2FD")]),
        ("ALIGN", (2,1), (3,-1), "CENTER"),
        ("ALIGN", (7,1), (7,-1), "RIGHT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
    ]))
    elems.append(table)
    doc.build(elems)

def export_history_excel(rows, save_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique"
    headers = ["Date", "Type", "Article", "Qté", "Destinataire", "Observation", "Stock après"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1976D2")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in rows:
        ws.append(list(r))
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = max(12, min(50, max_len+2))
    wb.save(save_path)

def export_history_pdf(rows, save_path, title="Historique Mouvements"):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    
    doc = SimpleDocTemplate(save_path, pagesize=landscape(A4), 
                           topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#1976D2")
    
    elems = [Paragraph(title, title_style), Spacer(1, 12)]
    
    # Définition des en-têtes (ici on les définit explicitement)
    headers = ["Date", "Type", "Article", "Qté", "Destinataire", "Observation", "Stock après"]
    
    data = [headers]  # ← maintenant headers existe
    for r in rows:
        data.append(list(r))
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1976D2")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor("#E3F2FD")]),
        ("ALIGN", (3,1), (3,-1), "CENTER"),
        ("ALIGN", (6,1), (6,-1), "RIGHT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
    ]))
    
    elems.append(table)
    doc.build(elems)